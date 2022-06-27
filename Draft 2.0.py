#This project was designed and created
#    by CDT Aaron W. Calhoun class of 2024 (H-2).

#Draft:2.0
#Started: Febuary 2021
#Draft Complete: August 25, 2021 
#Email: aaron.calhoun@westpoint.edu


#This program was designed to enable 
#   the user to find an optimal characteristic 
#   for their desired characteristics using a GUI.

#This program would not would not be possible without 
#   the help and guidance of Dr. Kenith Allen (United States Military Acadmey at West Point),
#   Rodney Keith (Sandia National Laboratories),
#   Darryn Fleming (Sandia National Laboratories),
#   and Logan Rapp  (Sandia National Laboratories).

# Things to note about the python code: 
# 1 You will need to restate the file address for within the python code so that it knows where to find the database.
# 2 You will need to import the necessary packages in-order to run the program (anaconda should come installed with most of them).
# 3 The qualitative section of the first tab (the drop down menus) DO NOT work … yet (I will resend this once I get this to work).
# 4 I was unable to get the scrollbar to work on the third tab which can make it hard to see all the data (especially on a lap top).
# 5 The sub-scores below each characteristics shows the percent deviation between the user input and the actual value for a given reactor using the input value from the user as its basis (0.01 = %1 difference, 1.0 = 100% difference).
# 6 The total scores in the third tab (after pressing the “calculate button”) is the final transformation including the weights of the user. (The lowest score will be the most similar to the users desired reactor)

    
from openpyxl import load_workbook 
import pandas as pd
from heapq import nlargest
import tkinter as tk
from tkinter import ttk
from tkinter import *
from pandastable import Table


#this program uses both openpyxl and pandas

#PLEASE RESTATE THE FILE ADDRESS FOR THE EXCEL SHEET
file="Database.xlsx"


workbook = load_workbook(filename=file)

workbook.sheetnames

['Sheet 1']

sheet = workbook.active

df=pd.read_excel(file)


#number of rows and columns
C = sheet.max_column
R = sheet.max_row

#list of characteristics within excel sheet
col=df.columns
length = len(col)


# used to determine, later, if a list is composed of numbers (quantitative) or letters (qualitative)
def is_number(n):
    try:
        float(n)
    except ValueError:
        return False
    return True
 


#list for quanitative and qualitative characteristics
qual=[]
quant=[]

index = -1

for x in range(0,C):
    index = index + 1

    for y in range(0,R):
        cell1 = str(df.iat[y,index])

        if str(cell1) == str("TBD") or str(cell1) == str("X"):
            continue

        else:
            if is_number(cell1) == True:
                quant.append(str(col[x]))

            else:
                qual.append(str(col[x]))

            break

   
#lists of lists for each of the qual characteristics 
qual_rep_list=[]
index = 0
for i in range(0,C):
    qual_rep=[]
    index = index + 1
    
    

    if col[i] in qual:
        for i in range(1,R-1):
            cell_rep = str(df.iat[i,index])
            tf = cell_rep in qual_rep
            
            if tf == False and is_number(cell_rep) == False:
                qual_rep.append(cell_rep)
                
        qual_rep_list.append(qual_rep)


    

#reactor names list
RN = [] 
for x in range(0,R-1):
    RN.append(str(df.iat[x,0]))



class app:
    def __init__(self,window):
        
        #creating tabs and window
        self.window = window

        self.window.title("Nuclear Reactors: Dataframe Analysis")

        tabControl = ttk.Notebook(window)
        

        self.tab1 = ttk.Frame(tabControl)

        self.tab2 = ttk.Frame(tabControl)

        self.tab3 = ttk.Frame(tabControl)

   

        tabControl.add(self.tab1, text ='Data Analysis Input')

        tabControl.add(self.tab2, text ='Input Reactor & Characteristics')

        tabControl.add(self.tab3, text ='Dataframe & Data Analysis Output')

        tabControl.grid()




        self.char_name  = []

        self.char_input = []
        
        self.reactor_input_value = []
        
        self.reactor_input_char_name = []

       
        #creating input boxes for name and reactor characteristics on tab 2
        c = -1
        for i in range(0,4):
            c=c+1
            
            self.reactor_input_char_name.append(tk.Text(self.tab2, height=1,width = 10))

            input_reactor = self.reactor_input_char_name[i]
            
            input_reactor.grid(row=len(col)+5+c,column = 1)

            
            self.reactor_input_value.append(tk.Text(self.tab2, height=1,width = 10))

            input_reactor1 = self.reactor_input_value[i]

            input_reactor1.grid(row=len(col)+5+c,column = 2)


    
        # creating labels and input boxes for tab 2 for new reactor 
        for i in range(0,len(col)):

            self.char_name.append(tk.Label(self.tab2,text=str(col[i])))

            self.char_input.append(tk.Text(self.tab2, height=1,width = 10))

           

            label_name=self.char_name[i]

            label_name.grid(row=i+3,column=0)

           

            input_char=self.char_input[i]

            input_char.grid(row=i+3,column=1)

        
        
        
        # constant labels for tab 1 
        
        tk.Label(self.tab1, text="Characteristic:").grid(row=0,column=0)

        tk.Label(self.tab1, text="Value:").grid(row=0,column=1)

        tk.Label(self.tab1, text="Weight:").grid(row=0,column=2)

        tk.Label(self.tab1, text="Characteristic:").grid(row=0,column=4)

        tk.Label(self.tab1, text="Value:").grid(row=0,column=5)

        tk.Label(self.tab1, text="Weight:").grid(row=0,column=6)
        
        
        
        # constant labels for tab 2
        
        tk.Label(self.tab2, text="Value for Reactor:").grid(row=2,column=1)
        
        tk.Label(self.tab2, text="Please input the characteristics for you reactor below.").grid(row=0,column=0)

        tk.Label(self.tab2, text="If you do not know a characteristic type 'TBD'. \n If a characteristic is not applicable type 'X'.").grid(row=1,column=0)

        tk.Label(self.tab2, text="Input a New Characteristic:").grid(row=len(col)+3,column=0)

        tk.Label(self.tab2, text="Characteristic:").grid(row=len(col)+4,column=1)

        tk.Label(self.tab2, text="Value for Reactor:").grid(row=len(col)+4,column=2)

        tk.Label(self.tab2, text="New Characteristic 1:").grid(row=len(col)+5,column=0)

        tk.Label(self.tab2, text="New Characteristic 2:").grid(row=len(col)+6,column=0)

        tk.Label(self.tab2, text="New Characteristic 3:").grid(row=len(col)+7,column=0)

        tk.Label(self.tab2, text="New Characteristic 4:").grid(row=len(col)+8,column=0)

       
       # list of weights and desired characteristics from the user 
        self.text_obj_list=[] 

        self.text_obj_list1=[]

       

        self.text_obj_list2=[]

        self.text_obj_list3=[]

       
        self.label_list=[]

        # adding the the labels and text boxes to tab 1
        for i in range(0,len(quant)):

            self.text_obj_list.append(tk.Text(self.tab1,height=0.2,width = 7)) #column one of four text boxes

            self.text_obj_list1.append(tk.Text(self.tab1,height=0.2,width = 7)) #column two of four text boxes

            self.text_obj_list2.append(tk.Text(self.tab1,height=0.2,width = 7)) #column three of four text boxes

            self.text_obj_list3.append(tk.Text(self.tab1,height=0.2,width = 7)) #column four of four text boxes

            self.label_list.append(tk.Label(self.tab1, text=str(quant[i])))
    
        
        # making user inputs into a list from tab 1 
        for i in range(0,len(self.label_list),2):
            
        
            text_obj=self.text_obj_list[i]
            text_obj1=self.text_obj_list1[i]
            text_obj2=self.text_obj_list2[i]
            text_obj3=self.text_obj_list3[i]
            
            label_obj=self.label_list[i]
            
            label_obj.grid(row=i+1,column=0)
            
            text_obj.grid(row=i+1,column=1)
            text_obj1.grid(row=i+1,column=2)
            
            if i+1 < int(len(quant)):
                
                label1_obj=self.label_list[i+1]
                label1_obj.grid(row=i+1,column=4)
                
                text_obj2.grid(row=i+1,column=5)
                text_obj3.grid(row=i+1,column=6)

   

         

        label=tk.Label(self.tab1,text="Check any acceptabe replacement characteristics")

        label.grid(columnspan=4)
       

        self.check_boxes=[] #list for the check boxes and menue boxes

        self.check_box_labels=[] # list of the lables for the boxes 

        self.check_box_qual_weight=[] # list of the input weights by the user 
        
        
        # creates both the titles and weight inputs for the qual 
        for i in range(1,len(qual)):
            self.check_box_qual_weight.append(tk.Text(self.tab1,height=0.2,width = 7))
            
            input_weight=self.check_box_qual_weight[i-1]
            
            input_weight.grid(row=len(quant)+i+4,column=2)


            self.check_box_labels.append(tk.Label(self.tab1, text=str(qual[i])))

            check_box_l=self.check_box_labels[i-1]

            check_box_l.grid(row=len(quant)+i+4,column=0)
            
        
        # creating menubuttons for tab1 based on the qual list
        for i in range(0,len(qual_rep_list)-1):
            self.mb =  Menubutton (self.tab1, text="Select", relief=RAISED)
            self.mb.menu  =  Menu ( self.mb, tearoff = 0 )
            self.mb["menu"]  =  self.mb.menu
            
            
            for j in range(0,len(qual_rep_list[i])):
                self.Item = IntVar() # this is the specfifc item being selected within the tab
                self.mb.menu.add_checkbutton(label=str(qual_rep_list[i][j]), variable=self.Item) #command = add_qual(i,j))?????
                
            else:
                self.check_boxes.append(self.mb)
                

        for i in range(0,len(self.check_boxes)):
            check_boxx=self.check_boxes[i]
            check_boxx.grid(row=len(quant)+i+5,column=1)
             
             
        # creating a list of of qual weights from the user
        list_of_qual_w = []
        for i in range(0, len(qual)-1):
            if str(self.check_box_qual_weight[i].get("1.0",tk.END)) == "\n":
                list_of_qual_w.append(0)
            else:
                list_of_qual_w.append(self.check_box_qual_weight[i].get("1.0",tk.END).translate({ord('\n'): None}))
        
        #converting user qual weight input into a list of int
        for i in range(0, len(list_of_qual_w)-1):
            list_of_qual_w[i]=float(list_of_qual_w[i])
        
        #button for tab 1 to run calculations and rewrite tab 3 
        def button_push():
            
           # creating lables for sub-score dataframe
           DataCalc = {}
           RN.insert(0,"") # needed for first row of description lables 
           RN.append("") 
           DataCalc["Reactor"] = RN
           
           #creating dataframe 
           for i in range(len(col)-1):
               DataCalc[str(col[i+1])] = None
               
           DataCalc["Total Scores (w/ Weight)"] = None 
           
           df_Sub = pd.DataFrame(DataCalc)
          
           
           Wsum = float(0)  #weight totals
           GV = [] #given values from user (desired reactor)
           W = [] #weights given from user
           
           #creating list for W and GV
           for i in range(0,len(self.text_obj_list),2):
                GV.append(float(self.text_obj_list[i].get("1.0","end-1c")))
                W.append(float(self.text_obj_list1[i].get("1.0","end-1c")))
                
                if i+1 < int(len(quant)):
                    GV.append(float(self.text_obj_list2[i].get("1.0","end-1c")))
                    W.append(float(self.text_obj_list3[i].get("1.0","end-1c")))
           Wsum = sum(W)
           
           #populated the data frame with percent difference between input and databse
           count = - 1
           for cc in range(1,len(col)):
               
               if str(col[cc]) in quant:
                    count = count + 1
                
               for rr in range(2,R+1):
                   
                   cell2 = str(df.at[rr-2,str(col[cc])])
                   
                   if is_number(cell2) == False:  
                       if cell2 == str(GV[1]): # GV[1] here is currently a stand in for a qualitative characteristic
                           S=1
                       
                       else:
                           S = 0 
                       
                       df_Sub.at[rr-1,str(col[cc])] = S
                      
                   else:
                        if is_number(cell2) == True: 
                            AV= float(cell2) #actual value of a given reactor (value from dataframe (real value))
                            df_Sub.at[rr-1,str(col[cc])] = round((((pow(pow((float(GV[count])-float(AV)),2.0),0.5))/float(GV[count]))),3)
                            
                        else:
                            AV = 0 
                            df_Sub.at[rr-1,str(col[cc])] = AV
                    
                        
           # populates the total score column and applies weight transformation  
           for i in range(1,R):
                Sum = 0
                count = -1 
                count = count + 1
                for j in range(1,C-1):
                    
                    cell = float(df_Sub.iat[i,j])*float(W[count])/float(Wsum)
                    Sum = Sum + cell
                df_Sub.at[i,"Total Scores (w/ Weight)"] = str(round(Sum,3))
           
           #creates new tab 3 and populates it with values for the dataframe of calculations
           self.tab3.destroy()
           self.tab3 = ttk.Frame(tabControl)
           tabControl.add(self.tab3, text ='Dataframe & Data Analysis Output')
           
           for cc in range(len(col)):
                 for rr in range(R):
                    tk.Label(self.tab3, text=str(col[cc])).grid(row=1,column=cc)
                    tk.Label(self.tab3, text=str(df_Sub.at[rr,str(col[cc])])).grid(row=rr+1,column=cc)
                    
           for rr in range(R):
                    tk.Label(self.tab3, text=str("Total Scores (w/ Weight)")).grid(row=1,column=C)
                    tk.Label(self.tab3, text=str(df_Sub.at[rr,"Total Scores (w/ Weight)"])).grid(row=rr+1,column=C)
                              
         
    
        button=tk.Button(self.tab1,text="Calculate",command=button_push)
        button.grid(columnspan=2)
 

        # button for tab 2 to populate the data base in the excel sheet     
        def button_push2():
                
            def is_number(n):
                try:
                    float(n)
                except ValueError:
                    return False
                return True
            
            
            #data base files and input files for user edits
            
            #importing database sheet
            file="C:\\Users\\Aaron.Calhoun\\OneDrive - West Point\\All\\N_Python\\total.xlsx"
            workbook = load_workbook(filename=file)
            workbook.sheetnames
            ['Sheet 1']
            sheet = workbook.active
            
            
            df=pd.read_excel(file)
            
            #number of rows and columns
            C = sheet.max_column
            R = sheet.max_row
            
            col=df.columns
            length = len(col)
        
            for x in range(1,C+1):
                if C > 26: 
                    for i in C: #not correct come back to
                        sheet[str(chr(ord('@')+x+1))+str(chr(ord('@')+i+1))+str(R+1)]= str(self.char_input[x-1].get("1.0", "end-1c"))
    
                else:
                    sheet[str(chr(ord('@')+x))+str(R+1)]= str(self.char_input[x-1].get("1.0", "end-1c")) #str(chr(ord('@') converts a number to its alphabetic equal (A=1, B=2, C=3 etc.)

            R = sheet.max_row 
            for x in range(0,len(self.reactor_input_value)): # fills the rest with TBD b/c they are unknown
            
                C = sheet.max_column
                for r in range(2,R): #prints out the title of the new charcteristic and value
                
                    if self.reactor_input_char_name[x].get("1.0", "end-1c") == "":
                        break
                    
                    if C > 26: 
                        sheet["A"+str(chr(ord('@')+C+1))+str(r)] = "TBD"
                    else:
                        sheet[str(chr(ord('@')+C+1))+str(r)] = "TBD"
                     
                    if C > 26: #if the excel sheet goes to AA, AB, AC ... ect.
                        sheet["A"+str(chr(ord('@')+C+1))+str(x)] = "TBD"  
                else:
                    if self.reactor_input_char_name[x].get("1.0", "end-1c") == "":
                        break
                    sheet[str(chr(ord('@')+C+1))+str(1)]= str(self.reactor_input_char_name[x].get("1.0", "end-1c")) 
                    sheet[str(chr(ord('@')+C+1))+str(R)]= str(self.reactor_input_value[x].get("1.0", "end-1c"))
            
            
            List_add=[]
            for i in range(0,len(self.char_input)):
                List_add.append(str(self.char_input[i].get("1.0", "end-1c"))) 
    

            df=pd.read_excel(file)
            col=df.columns
            length = len(col)
                             
            
            
            sheet.title
            'Sheet1'
            
            workbook.save(filename = file)
            
        
        button1=tk.Button(self.tab2,text="Add Reactor",command=button_push2)
        button1.grid(columnspan=2)
        
        C = sheet.max_column
        R = sheet.max_row
        
        # populates tab 3 with the data frame and column characteristics names 
        for cc in range(len(col)-1):
               tk.Label(self.tab3, text=str(col[cc])).grid(row=1,column=cc)
       
        for i in range(1,R):
            for j in range(1,C):
                tk.Label(self.tab3, text=str(df.iat[i-1,j-1])).grid(row=i+1,column=j-1)

        # unabel to get the scroll bar to work on tab 3 and unable
        
        # tk.Scrollbar(self.tab3).grid(row=0, column = C+C)
        # scrollbar = Scrollbar(self.tab3)
        # scrollbar.grid(row=0, column = C+C)
        # scrollbar.config(window)
        
window=tk.Tk()
start=app(window)
window.mainloop() 
