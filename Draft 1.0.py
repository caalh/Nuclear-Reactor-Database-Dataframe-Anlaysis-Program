#This project was designed and created
#    by CDT Aaron W. Calhoun class of 2024 (H-2).

#Draft:1.0
#Started: Febuary 2021
#Draft Complete: August 25, 2021 
#Email: aaron.calhoun@westpoint.edu


#This program was designed to enable 
#   the user to find an optimal characteristic 
#   for their desired characteristics.

#This program would not would not be possible without 
#   the help and guidance of Dr. Kenneth Allen (United States Military Acadmey at West Point).
    


from openpyxl import load_workbook 
import pandas as pd
from heapq import nlargest


 

def is_number(n):
    try:
        float(n)
    except ValueError:
        return False
    return True


#data base files and input files for user edits

#importing database sheet

#PLEASE RESTATE THE FILE ADDRESS FOR THE EXCEL SHEET
file="C:\\Users\\Aaron.Calhoun\\OneDrive - West Point\\All\\N_Python\\total.xlsx" 
workbook = load_workbook(filename=file)
workbook.sheetnames
['Sheet 1']
sheet = workbook.active


df=pd.read_excel(file)
print(df)

#number of rows and columns
C = sheet.max_column
R = sheet.max_row

col=df.columns
length = len(col)


#options for user
yn=input("Would you like to input a reactor (1) or compare characteristics (2)")

#input reactor
if yn == str(1):
    print("If you do not know a characteristic type 'TBD'. If a characteristic is not applicable type 'x'.")
    print("")
    titles=[]
    for i in range(length):
        titles.append(input(col[i]+":"))

    for x in range(1,C+1):
        sheet[str(chr(ord('@')+x))+str(R+1)]= titles[x-1] # str(chr(ord('@') converts a number to its alphabetic equal (A=1, B=2, C=3 etc.)


    
    YorN = str(input("Would you like to add a new characteristic? (y/n)")).upper()
    if YorN == "Y":
        
        C = sheet.max_column
        R = sheet.max_row
        
        Name = input("What is the characteristic?")
        RValue = input("What is the value of your reactor for that characteristic?")
        
        if C > 26: 
            sheet[str(chr(ord('@')+C+1))+str(1)] = str(Name)
            sheet[str(chr(ord('@')+C+1))+str(chr(ord('@'+1+C)))+str(R)] = str(RValue)  
        else:
            sheet[str(chr(ord('@')+C+1))+str(1)] = str(Name)
            sheet[str(chr(ord('@')+C+1))+str(R)] = str(RValue)        

       
        for x in range(2,R):
            if C > 26: 
                sheet["A"+str(chr(ord('@')+C+1))+str(x)] = "TBD"  
            else:
                sheet[str(chr(ord('@')+C+1))+str(x)] = "TBD" # LIMITATION: this is unbable to past "AZ" as BA, BB, BC .... does not work
            

        
        
    df=pd.read_excel(file)
    col=df.columns
    length = len(col)
                     
    
    
    sheet.title
    'Sheet1'
    
    workbook.save(filename = file)
    
elif str(yn) != str(2):
    print("Input Error.")
  
#comparison
else:
    qual=[]
    quant=[]
    
    index = -1
    for x in range(0,C):
        
        index = index + 1
        
        for y in range(0,R):
            cell1 = str(df.iat[y,index]) 
            
            if str(cell1) == str("TBD") or str(cell1) == str("X"):
                continue
                
            else:
                if is_number(cell1) == True:
                    quant.append(str(col[x]))
                        
                else:
                    qual.append(str(col[x]))
                    
                break


    A = [] # values 
    AA = [] # weights for quant
    RN = [] #reactor name
    RV = [] #replacement or null values for qual
    NV = [] # null values for quant
    WQual = []
    #asking values and weights and getting reactor names 
    for x in range(0,R-1):
        RN.append(str(df.iat[x,0]))

    for x in range(len(quant)):
        A.append(input(quant[x]+":").upper())
        AA.append(int(input(quant[x]+" weight (number):")))
    
    for x in range(len(qual)-1):
        WQual.append(int(input(qual[x+1] + " weight:")))
        RV.append(input("Name a suitable replacement for " + qual[x+1] + " of your reactor:"))

   
 #inputing values into sheet

    # creating lables for sub-score dataframe
    DataCalc = {}
    DataCalc["Reactor"] = RN
    
    for i in range(len(col)-1):
        DataCalc[str(col[i+1])] = None
        
    df_Sub = pd.DataFrame(DataCalc)


    #weight totals
    Wsum = float(sum(WQual)+sum(AA))
    GV = [] #given values from user 
    W = []#weights given from user
    #loop for creating subscores
    GV = A
    W = AA
    # for x in range(len(GV)):
        #con = user input of acceptable solution 
        #completes the transformation for sub-totals

    for cc in range(0,len(GV)):
        for rr in range(2,R+1):
            
            cell2 = str(df.at[rr-2,str(quant[cc])])
        
            if is_number(GV[cc]) == False:                 
                if str(GV[cc]) == "TBD":
                  S = 0 #Sub-score
                  
                elif str(GV[cc]) in NV:
                  S = 1
                  
                else:
                  S = 0 

                df_Sub.at[cc,str(quant[rr-2])] = S
                
            else:
                if is_number(cell2) == True: 
                    AV= float(cell2) #actual value of a given reactor
                    df_Sub.at[rr-2,str(quant[cc])] = (1/((pow(pow((float(GV[cc])-float(AV)),2.0),0.5))/float(AV)))*float(W[cc])/float(Wsum)
                    
                else:
                    AV = 0 
                    df_Sub.at[rr-2,str(quant[cc])] = AV


        #completes the transformation for sub-totals
    for cc in range(1,len(RV)+1):
        for rr in range(2,R+1):
            
            cell2 = str(df.at[rr-2,str(qual[cc])])
            
            if is_number(cell2) == False:                
                if str(cell2) == "TBD":
                  S = 0 #Sub-score
                  
                elif str(RV[cc-1]) == cell2:
                  S = 1
                  
                else:
                    S = 0
                df_Sub.at[rr-2,str(qual[cc])] = S
                
            else:
                AV= float(cell2) #actual value of a given reactor
                df_Sub.at[rr-2,str(qual[cc])] = (1/((pow(pow((float(RV[cc-1])-float(AV)),2.0),0.5))/float(RV[cc-1])))*float(W[cc])/float(Wsum)
            
                # AV = 0 
                # df_Sub.at[rr-2,str(qual[cc])] = AV


    #calculates final total for reactors and finds heightest value
    
    print(df_Sub)
    SOL = []
    for y in range(2,R-1):
        Numb = float(0)
        for x in range(2,C):
            Numb = Numb + float(df_Sub.at[y,str(col[x])])
        SOL.append(Numb*100)
        
    
    
    dc = {}
    for i in range(0,R-3):
        dc[str(df_Sub.at[i,str("Reactor")])] = SOL[i]
    
    
    ThreeHighest = nlargest(R-1, dc, key = dc.get)
    for val in ThreeHighest:
        print(val, " : ", dc.get(val))